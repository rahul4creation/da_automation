import sys
from pathlib import Path

from openpyxl import load_workbook


MAX_SHEETS = 8
MAX_ROWS = 12
MAX_COLS = 12


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return text[:160]


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: excel-preview.py <xlsx-file>")

    workbook_path = Path(sys.argv[1])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    lines = [
        f"Excel workbook: {workbook_path.name}",
        f"Sheets: {', '.join(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:MAX_SHEETS]:
        sheet = workbook[sheet_name]
        lines.append("")
        lines.append(f"Sheet: {sheet_name}")
        lines.append(f"Used range estimate: {sheet.max_row} rows x {sheet.max_column} columns")

        row_count = 0
        for row in sheet.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True):
            values = [clean(value) for value in row]
            if any(values):
                lines.append(" | ".join(values))
                row_count += 1

        if row_count == 0:
            lines.append("(No non-empty preview rows found.)")

    print("\n".join(lines))


if __name__ == "__main__":
    main()
