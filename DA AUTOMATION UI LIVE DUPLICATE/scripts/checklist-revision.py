import argparse
import copy
import json
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def find_header(sheet):
    for row in range(1, min(sheet.max_row, 12) + 1):
        values = [clean(sheet.cell(row=row, column=col).value).lower() for col in range(1, sheet.max_column + 1)]
        serial_col = next((index + 1 for index, value in enumerate(values) if value in {"s.no", "s no", "sr no", "serial no"}), 1)
        point_col = next((index + 1 for index, value in enumerate(values) if "check" in value and "point" in value), 2)
        if serial_col and point_col and any(values):
            return row, serial_col, point_col
    return 1, 1, 2


def sheet_info(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        header_row, serial_col, point_col = find_header(sheet)
        point_count = 0
        for row in range(header_row + 1, sheet.max_row + 1):
            if clean(sheet.cell(row=row, column=point_col).value):
                point_count += 1
        sheets.append(
            {
                "name": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "headerRow": header_row,
                "serialColumn": serial_col,
                "pointColumn": point_col,
                "pointCount": point_count,
            }
        )
    return {"sheets": sheets}


def last_data_row(sheet, header_row, serial_col, point_col):
    last_row = header_row
    for row in range(header_row + 1, sheet.max_row + 1):
        if clean(sheet.cell(row=row, column=serial_col).value) or clean(sheet.cell(row=row, column=point_col).value):
            last_row = row
    return last_row


def first_data_row(sheet, header_row, serial_col, point_col):
    for row in range(header_row + 1, sheet.max_row + 1):
        if clean(sheet.cell(row=row, column=serial_col).value) or clean(sheet.cell(row=row, column=point_col).value):
            return row
    return header_row + 1


def json_cell(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return clean(value)


def sheet_grid(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f'Sheet "{sheet_name}" was not found in the checklist workbook.')

    sheet = workbook[sheet_name]
    header_row, serial_col, point_col = find_header(sheet)
    data_start = first_data_row(sheet, header_row, serial_col, point_col)
    data_end = last_data_row(sheet, header_row, serial_col, point_col)
    max_col = max(sheet.max_column, serial_col, point_col, 2)
    columns = []
    for col in range(1, max_col + 1):
        columns.append(clean(sheet.cell(row=header_row, column=col).value) or f"Column {col}")

    rows = []
    for row in range(data_start, data_end + 1):
        values = [json_cell(sheet.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
        if any(clean(value) for value in values):
            rows.append({"rowNumber": row, "values": values})

    return {
        "sheetName": sheet.title,
        "headerRow": header_row,
        "firstDataRow": data_start,
        "maxColumn": max_col,
        "serialColumn": serial_col,
        "pointColumn": point_col,
        "columns": columns,
        "rows": rows,
    }


def next_serial(sheet, header_row, serial_col):
    serial = 0
    for row in range(header_row + 1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=serial_col).value
        try:
            serial = max(serial, int(float(value)))
        except (TypeError, ValueError):
            continue
    return serial + 1


def copy_row_style(sheet, source_row, target_row):
    if source_row <= 0:
        return
    for col in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=col)
        target = sheet.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.border:
            target.border = copy.copy(source.border)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.font:
            target.font = copy.copy(source.font)


def append_points(source_path, output_path, sheet_name, points):
    workbook = load_workbook(source_path)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f'Sheet "{sheet_name}" was not found in the checklist workbook.')

    sheet = workbook[sheet_name]
    header_row, serial_col, point_col = find_header(sheet)
    append_row = last_data_row(sheet, header_row, serial_col, point_col) + 1
    serial = next_serial(sheet, header_row, serial_col)
    style_row = append_row - 1 if append_row > header_row else header_row
    added_count = 0

    for point in points:
        text = clean(point)
        if not text:
            continue
        copy_row_style(sheet, style_row, append_row)
        sheet.cell(row=append_row, column=serial_col).value = serial
        sheet.cell(row=append_row, column=point_col).value = text
        serial += 1
        append_row += 1
        added_count += 1

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {"savedPath": str(output), "sheetName": sheet_name, "addedCount": added_count}


def normalize_edit_rows(rows, max_col, serial_col):
    normalized = []
    for row in rows:
        values = row.get("values") if isinstance(row, dict) else row
        if not isinstance(values, list):
            continue
        padded = [clean(values[index]) if index < len(values) else "" for index in range(max_col)]
        content_values = [value for index, value in enumerate(padded, start=1) if index != serial_col]
        if any(content_values):
            normalized.append(padded)
    return normalized


def worksheet_text_value(value):
    text = clean(value)
    if not text:
        return None
    if text.startswith("="):
        return f"'{text}"
    return text


def save_sheet_revision(source_path, output_path, sheet_name, rows):
    workbook = load_workbook(source_path)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f'Sheet "{sheet_name}" was not found in the checklist workbook.')

    sheet = workbook[sheet_name]
    header_row, serial_col, point_col = find_header(sheet)
    data_start = first_data_row(sheet, header_row, serial_col, point_col)
    data_end = last_data_row(sheet, header_row, serial_col, point_col)
    max_col = max(sheet.max_column, serial_col, point_col, 2)
    normalized_rows = normalize_edit_rows(rows, max_col, serial_col)
    style_row = data_start if data_start <= data_end else header_row

    for row in range(data_start, max(data_end, data_start + len(normalized_rows) - 1) + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).value = None

    for row_index, row_values in enumerate(normalized_rows):
        target_row = data_start + row_index
        copy_row_style(sheet, min(style_row + row_index, data_end) if data_start <= data_end else style_row, target_row)
        for col in range(1, max_col + 1):
            if col == serial_col:
                sheet.cell(row=target_row, column=col).value = row_index + 1
            else:
                sheet.cell(row=target_row, column=col).value = worksheet_text_value(row_values[col - 1])

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {"savedPath": str(output), "sheetName": sheet_name, "updatedCount": len(normalized_rows)}


def main():
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect")
    inspect_parser.add_argument("workbook")

    sheet_parser = subparsers.add_parser("sheet")
    sheet_parser.add_argument("workbook")
    sheet_parser.add_argument("sheet")

    append_parser = subparsers.add_parser("append")
    append_parser.add_argument("source")
    append_parser.add_argument("output")
    append_parser.add_argument("sheet")
    append_parser.add_argument("points_json")

    save_sheet_parser = subparsers.add_parser("save-sheet")
    save_sheet_parser.add_argument("source")
    save_sheet_parser.add_argument("output")
    save_sheet_parser.add_argument("sheet")
    save_sheet_parser.add_argument("rows_json")

    args = parser.parse_args()
    if args.command == "inspect":
        print(json.dumps(sheet_info(Path(args.workbook))))
        return

    if args.command == "sheet":
        print(json.dumps(sheet_grid(Path(args.workbook), args.sheet)))
        return

    if args.command == "append":
        points = json.loads(args.points_json)
        if not isinstance(points, list) or not any(clean(point) for point in points):
            raise SystemExit("At least one checklist point is required.")
        print(json.dumps(append_points(Path(args.source), Path(args.output), args.sheet, points)))
        return

    rows = json.loads(args.rows_json)
    if not isinstance(rows, list):
        raise SystemExit("Checklist sheet rows are required.")
    print(json.dumps(save_sheet_revision(Path(args.source), Path(args.output), args.sheet, rows)))


if __name__ == "__main__":
    main()
